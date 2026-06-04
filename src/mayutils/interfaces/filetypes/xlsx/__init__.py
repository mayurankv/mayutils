"""
Expose XLSX workbook and sheet handles for tabular file IO.

A single ``.xlsx`` file bundles one or more sheets, which breaks the
"one file is one DataFrame" contract the rest of the
:class:`~mayutils.interfaces.filetypes.DataFile` hierarchy assumes.
The two classes here split the responsibilities cleanly: :class:`Xlsx`
is a workbook-level handle that enumerates sheets and returns mappings
of sheet name to DataFrame, and :class:`XlsxSheet` is the registered
:class:`DataFile` subclass that binds to a single ``(path, sheet)``
pair. The latter is what :meth:`DataFile.from_path` dispatches to when
the ``.xlsx`` suffix is encountered alongside a ``sheet=`` keyword.

See Also
--------
openpyxl.Workbook : Backing workbook object used to enumerate sheets.
openpyxl.worksheet.worksheet.Worksheet : Backing worksheet object used
    for row counting.
pandas.read_excel : Underlying pandas reader for XLSX sheets.
pandas.ExcelWriter : Underlying pandas writer for XLSX sheets.
mayutils.interfaces.filetypes.DataFile : Abstract base that
    :class:`XlsxSheet` implements.

Examples
--------
>>> import tempfile
>>> import pandas as pd
>>> from pathlib import Path
>>> from mayutils.interfaces.filetypes.xlsx import Xlsx, XlsxSheet
>>> with tempfile.TemporaryDirectory() as tmp:
...     path = Path(tmp) / "report.xlsx"
...     pd.DataFrame({"a": [1, 2, 3]}).to_excel(path, sheet_name="Summary", index=False)
...     workbook = Xlsx(path)
...     names = workbook.sheet_names
...     sheet = XlsxSheet(path, sheet="Summary")
...     frame = sheet.read()
>>> names
['Summary']
>>> frame.shape
(3, 1)
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any, ClassVar, Self, cast

from mayutils.core.extras import may_require_extras
from mayutils.interfaces.filetypes import DataFile
from mayutils.objects.dataframes.backends import Backend, DataFrames, default_backend

with may_require_extras():
    import openpyxl
    import pandas as pd
    import polars as pl
    from pandas import ExcelWriter
    from xlsxwriter import Workbook

if TYPE_CHECKING:
    from collections.abc import Iterator


DEFAULT_SCHEMA_SAMPLE_ROWS = 1000


class Xlsx[DataFrameType: DataFrames = pd.DataFrame]:
    """
    Represent a workbook-level handle over an ``.xlsx`` file.

    Holds a path to an Excel workbook and exposes helpers for
    enumerating sheet names, constructing per-sheet
    :class:`XlsxSheet` views, and reading or writing the whole
    workbook as a mapping of sheet name to DataFrame. Intentionally
    not a :class:`DataFile` subclass because its read and write
    contract returns and accepts a ``dict[str, ??]`` rather
    than a single DataFrame, which does not fit the one-file
    one-frame abstraction used elsewhere.

    Parameters
    ----------
    path
        Filesystem location of the workbook. The suffix must be
        ``.xlsx`` (case-insensitive).
    backend
        Default DataFrame backend for reads, writes, and for any
        :class:`XlsxSheet` handles spawned via :meth:`sheet`.

    Attributes
    ----------
    path
        Resolved path to the workbook file.
    backend
        Default backend captured at construction.

    Raises
    ------
    ValueError
        If ``path`` does not have a ``.xlsx`` suffix.

    See Also
    --------
    XlsxSheet : Single-sheet :class:`DataFile` counterpart.
    openpyxl.Workbook : Backing workbook used to enumerate sheets.
    mayutils.interfaces.filetypes.DataFile : Abstract tabular file
        interface.

    Examples
    --------
    >>> import tempfile
    >>> import pandas as pd
    >>> from pathlib import Path
    >>> with tempfile.TemporaryDirectory() as tmp:
    ...     path = Path(tmp) / "report.xlsx"
    ...     pd.DataFrame({"a": [1, 2]}).to_excel(path, sheet_name="Summary", index=False)
    ...     workbook = Xlsx(path)
    ...     backend = workbook.backend.name
    ...     names = workbook.sheet_names
    >>> backend
    'pandas'
    >>> names
    ['Summary']
    """

    suffix: ClassVar[str] = ".xlsx"

    @staticmethod
    def column_to_excel(
        column: int,
        /,
    ) -> str:
        """
        Translate a one-indexed column ordinal into spreadsheet letter notation.

        Applies the bijective base-26 encoding used by spreadsheets to
        map positive integers onto the sequence ``A``, ``B``, ...,
        ``Z``, ``AA``, ``AB``, ..., ``ZZ``, ``AAA``, ... This is the
        inverse of the letter-based column addressing used by
        Excel-compatible exporters across the library. Lives on
        :class:`Xlsx` rather than as a module-level helper so callers
        can reach it via the workbook type without importing a
        separate name.

        Parameters
        ----------
        column
            The one-indexed position of the column, where ``1``
            denotes the leftmost column. Values of ``0`` or below
            yield an empty string because the base-26 loop terminates
            immediately, so callers are expected to pass strictly
            positive integers.

        Returns
        -------
            The uppercase letter sequence that identifies the column
            in spreadsheet notation, composed solely of ASCII
            characters ``A``-``Z``.

        See Also
        --------
        XlsxSheet.excel_range : Consumer that turns the letters into a
            full ``Sheet!A1:Z10`` range reference.
        openpyxl.utils.cell.get_column_letter : Equivalent helper from
            openpyxl.

        Examples
        --------
        >>> Xlsx.column_to_excel(1)
        'A'
        >>> Xlsx.column_to_excel(27)
        'AA'
        >>> Xlsx.column_to_excel(702)
        'ZZ'
        """
        result: list[str] = []

        while column > 0:
            column, rem = divmod(column - 1, 26)
            result.append(chr(65 + rem))

        return "".join(reversed(result))

    def __init__(
        self,
        path: Path | str,
        /,
        *,
        backend: Backend[DataFrameType] | None = None,
    ) -> None:
        """
        Bind the workbook handle to ``path`` with a default DataFrame backend.

        Normalises ``path`` to a :class:`pathlib.Path`, stores the
        default backend for later reads and writes, and validates the
        file suffix so downstream openpyxl and pandas calls are
        guaranteed to see an Excel file. No IO happens at construction
        time; the workbook is only touched when :meth:`exists`,
        :meth:`sheet_names`, :meth:`read`, or :meth:`write` runs.

        Parameters
        ----------
        path
            Workbook file path.
        backend
            Default DataFrame backend for workbook-level reads and
            writes and for spawned :class:`XlsxSheet` instances.

        Raises
        ------
        ValueError
            If ``path`` has a suffix other than ``.xlsx``.

        See Also
        --------
        Xlsx.sheet : Spawn a per-sheet handle using this backend.
        XlsxSheet.__init__ : Single-sheet counterpart.

        Examples
        --------
        >>> workbook = Xlsx("report.xlsx", backend="polars")
        >>> workbook.backend
        'polars'
        """
        self.path = Path(path)
        self.backend = backend if backend is not None else cast("Backend[DataFrameType]", default_backend())

        if self.path.suffix.lower() != self.suffix:
            msg = f"Xlsx expects a '{self.suffix}' file; got '{self.path.suffix}'."
            raise ValueError(msg)

    def _identity(
        self,
    ) -> str:
        """
        Return the identity header for this workbook handle.

        Used by :meth:`__repr__` and :meth:`_repr_html_` as a short,
        IO-free description of the handle. Keeping it as a separate
        method mirrors the pattern used by
        :class:`~mayutils.interfaces.filetypes.DataFile` so repr
        overrides can cheaply compose the header before attempting to
        load sheet previews.

        Returns
        -------
            ``"Xlsx(<path>, backend=<backend>)"``.

        See Also
        --------
        Xlsx.__repr__ : Consumer of this header.
        XlsxSheet._identity : Single-sheet analogue.

        Examples
        --------
        >>> Xlsx("report.xlsx")._identity()
        "Xlsx(report.xlsx, backend='pandas')"
        """
        return f"Xlsx({self.path!s}, backend={self.backend.name!r})"

    def __repr__(
        self,
    ) -> str:
        """
        Return an identity header followed by a preview of each sheet.

        When the workbook exists, one preview block per sheet is
        included using the sheet's :class:`XlsxSheet` repr. Failures
        fall back to the identity line alone so ``repr`` never
        raises, matching the defensive contract used by the rest of
        the :class:`DataFile` hierarchy.

        Returns
        -------
            Multi-line string; a single line when the workbook is
            unavailable.

        See Also
        --------
        Xlsx._repr_html_ : Notebook-friendly counterpart.
        XlsxSheet.__repr__ : Per-sheet preview used here.

        Examples
        --------
        >>> import tempfile
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "missing.xlsx"
        ...     header = repr(Xlsx(path))
        >>> header.startswith("Xlsx(")
        True
        >>> header.endswith("backend='pandas')")
        True
        """
        header = self._identity()
        if not self.exists():
            return header
        try:
            names = self.sheet_names
        except Exception:  # noqa: BLE001
            return header
        if not names:
            return f"{header}\n(no sheets)"

        sections = "\n\n".join(f"[{name}]\n{self.sheet(name)!r}" for name in names)

        return f"{header}\n{sections}"

    def _repr_html_(
        self,
    ) -> str:
        """
        Render the workbook as a concatenation of per-sheet HTML previews.

        Used by Jupyter / IPython to show a rich workbook preview
        inline. Each sheet's HTML is delegated to
        :meth:`XlsxSheet._repr_html_`, with a headline built from the
        sheet name so viewers can quickly scan multiple tabs in one
        notebook cell. Failures collapse to the identity header so
        cell output never raises.

        Returns
        -------
            HTML string; falls back to the identity line on any
            error.

        See Also
        --------
        Xlsx.__repr__ : Plain-text counterpart.
        XlsxSheet._repr_html_ : Per-sheet HTML preview used here.

        Examples
        --------
        >>> import tempfile
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "missing.xlsx"
        ...     html = Xlsx(path)._repr_html_()
        >>> html.startswith("<code>Xlsx(")
        True
        >>> html.endswith(")</code>")
        True
        """
        header = f"<code>Xlsx({self.path!s})</code>"
        if not self.exists():
            return header
        try:
            names = self.sheet_names
        except Exception:  # noqa: BLE001
            return header
        if not names:
            return f"{header}<br/>(no sheets)"

        sections = "".join(f"<h4><code>{name}</code></h4>{self.sheet(name)._repr_html_()}" for name in names)  # pyright: ignore[reportPrivateUsage]

        return f"{header}{sections}"

    def exists(
        self,
    ) -> bool:
        """
        Return whether the workbook file exists on disk.

        Delegates to :meth:`pathlib.Path.is_file` so symlinks to
        regular files count as existing while directories and missing
        paths do not. Used by the repr helpers to decide whether to
        attempt sheet enumeration and by :meth:`_write` to pick
        between create and append modes.

        Returns
        -------
            ``True`` iff :attr:`path` resolves to a regular file.

        See Also
        --------
        pathlib.Path.is_file : Backing existence check.
        XlsxSheet._write : Consumer that branches on this flag.

        Examples
        --------
        >>> Xlsx("missing.xlsx").exists()
        False
        """
        return self.path.is_file()

    @property
    def sheet_names(
        self,
    ) -> list[str]:
        """
        Return the workbook's sheet names in declaration order.

        Opens the workbook in read-only, data-only mode via
        :func:`openpyxl.load_workbook` so the call is cheap and does
        not load cell formulae or styles. The workbook handle is
        closed explicitly in a ``finally`` block to free file
        descriptors even when the caller abandons the result.

        Returns
        -------
            Sheet names as reported by openpyxl's read-only workbook
            loader, preserving the author-defined ordering.

        See Also
        --------
        openpyxl.load_workbook : Backing loader used for enumeration.
        openpyxl.Workbook.sheetnames : Underlying attribute read here.
        Xlsx.sheet : Consumer that materialises one sheet at a time.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ...         pd.DataFrame({"a": [1]}).to_excel(writer, sheet_name="Summary", index=False)
        ...         pd.DataFrame({"b": [2]}).to_excel(writer, sheet_name="Details", index=False)
        ...     names = Xlsx(path).sheet_names
        >>> sorted(names)
        ['Details', 'Summary']
        """
        workbook = openpyxl.load_workbook(
            filename=self.path,
            read_only=True,
            data_only=True,
        )

        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def sheet(
        self,
        sheet: str,
        /,
    ) -> XlsxSheet[DataFrameType]:
        """
        Return an :class:`XlsxSheet` view bound to one sheet of the workbook.

        Forwards :attr:`path` and :attr:`backend` to the new handle so
        the per-sheet view inherits the workbook-level defaults. The
        returned handle satisfies the full :class:`DataFile`
        interface, meaning callers can treat the sheet as a
        first-class tabular file without going back through the
        workbook.

        Parameters
        ----------
        sheet
            Sheet name to scope the view to. Must already exist in
            the workbook for read operations; for write operations a
            new sheet with this name will be created or replaced.

        Returns
        -------
            A single-sheet handle that satisfies the full
            :class:`DataFile` interface.

        See Also
        --------
        XlsxSheet : Returned type.
        Xlsx.sheet_names : Enumeration helper used to discover valid
            ``name`` values.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     pd.DataFrame({"a": [1, 2]}).to_excel(path, sheet_name="Summary", index=False)
        ...     workbook = Xlsx(path)
        ...     sheet = workbook.sheet("Summary")
        ...     frame = sheet.read()
        >>> sheet.sheet
        'Summary'
        >>> frame.shape
        (2, 1)
        """
        return XlsxSheet(
            self.path,
            sheet=sheet,
            backend=self.backend,
        )

    def read(
        self,
        **kwargs: Any,  # noqa: ANN401
    ) -> dict[str, DataFrameType]:
        """
        Read every sheet in the workbook into a DataFrame mapping.

        Chooses the reader path based on the effective backend. Polars
        uses :func:`polars.read_excel` with ``sheet_id=0`` which
        returns all sheets via the calamine engine and avoids a pandas
        round-trip entirely. Pandas uses :func:`pandas.read_excel`
        with ``sheet_name=None`` and converts the returned
        :class:`collections.OrderedDict` into a plain ``dict`` to
        standardise the return type.

        Parameters
        ----------
        **kwargs
            Forwarded to :func:`pandas.read_excel` for the underlying
            read. The ``sheet_name=None`` option is set by this
            method.

        Returns
        -------
            Sheet name to DataFrame mapping. The polars branch reads
            via :func:`polars.read_excel` (calamine-backed) so
            pandas isn't touched; the pandas branch goes through
            :func:`pandas.read_excel`.

        See Also
        --------
        Xlsx.write : Inverse operation.
        pandas.read_excel : Backing pandas reader.
        polars.read_excel : Backing polars reader.
        XlsxSheet._read : Single-sheet counterpart.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     with pd.ExcelWriter(path, engine="openpyxl") as writer:
        ...         pd.DataFrame({"a": [1]}).to_excel(writer, sheet_name="Summary", index=False)
        ...         pd.DataFrame({"b": [2]}).to_excel(writer, sheet_name="Details", index=False)
        ...     frames = Xlsx(path).read()
        >>> sorted(frames)
        ['Details', 'Summary']
        >>> frames["Summary"].columns.tolist()
        ['a']
        """
        if self.backend.name == "polars":
            return pl.read_excel(  # pyright: ignore[reportUnknownVariableType]
                source=self.path,
                sheet_id=0,
                **kwargs,
            )

        return cast(
            "dict[str, DataFrameType]",
            pd.read_excel(  # pyright: ignore[reportUnknownMemberType]
                io=self.path,
                sheet_name=None,
                **kwargs,
            ),
        )

    def write(
        self,
        frames: dict[str, DataFrames],
        /,
        **kwargs: Any,  # noqa: ANN401
    ) -> Self:
        """
        Write a mapping of sheet name to DataFrame into the workbook.

        Any pre-existing workbook at :attr:`path` is replaced. The
        write path is chosen per call: when every frame is a
        :class:`polars.DataFrame`, each sheet is written natively via
        :meth:`polars.DataFrame.write_excel` against a shared
        :class:`xlsxwriter.Workbook`, avoiding a pandas round-trip.
        Otherwise all frames are coerced to pandas and written
        through :class:`pandas.ExcelWriter` so mixed-backend dicts
        keep working.

        Parameters
        ----------
        frames
            Sheet name to DataFrame mapping. Empty mappings raise a
            :class:`ValueError` because a workbook must contain at
            least one sheet.
        **kwargs
            Forwarded to the backend writer. For the polars path,
            they go to :meth:`polars.DataFrame.write_excel`; for the
            pandas path, to :meth:`pandas.DataFrame.to_excel`.

        Returns
        -------
            The current handle, for fluent chaining.

        Raises
        ------
        ValueError
            If ``frames`` is empty.

        See Also
        --------
        Xlsx.read : Inverse operation.
        pandas.ExcelWriter : Backing pandas writer.
        polars.DataFrame.write_excel : Backing polars writer.
        XlsxSheet._write : Single-sheet counterpart.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "out.xlsx"
        ...     frames = {"Summary": pd.DataFrame({"a": [1, 2]})}
        ...     handle = Xlsx(path).write(frames)
        ...     readback = Xlsx(path).read()
        >>> isinstance(handle, Xlsx)
        True
        >>> sorted(readback)
        ['Summary']
        >>> readback["Summary"].shape
        (2, 1)
        """
        if not frames:
            msg = "Cannot write an XLSX workbook with zero sheets."
            raise ValueError(msg)

        if all(isinstance(frame, pl.DataFrame) for frame in frames.values()):
            with Workbook(filename=self.path) as workbook:  # pyright: ignore[reportUnknownVariableType]
                for sheet_name, frame in frames.items():
                    cast("pl.DataFrame", frame).write_excel(
                        workbook=workbook,  # pyright: ignore[reportUnknownArgumentType]
                        worksheet=sheet_name,
                        **kwargs,
                    )

            return self

        index_kwarg = kwargs.pop("index", False)
        with ExcelWriter(path=self.path, engine="openpyxl") as writer:
            for sheet_name, frame in frames.items():
                if isinstance(frame, pl.DataFrame):  # noqa: SIM108
                    pandas_frame = frame.to_pandas()
                else:
                    pandas_frame = frame

                pandas_frame.to_excel(  # pyright: ignore[reportUnknownMemberType]
                    excel_writer=writer,
                    sheet_name=sheet_name,
                    index=index_kwarg,
                    **kwargs,
                )

        return self


class XlsxSheet[DataFrameType: DataFrames = pd.DataFrame](DataFile[DataFrameType]):
    """
    Handle a single sheet within an ``.xlsx`` workbook as a :class:`DataFile`.

    Extends :class:`DataFile` so callers can treat a sheet as a
    first-class tabular file: read or write DataFrames, introspect
    the schema, count rows, and stream chunks. Reads go through
    :func:`pandas.read_excel`; writes use
    :class:`pandas.ExcelWriter` with ``if_sheet_exists="replace"`` so
    an existing workbook gets its target sheet overwritten in place
    rather than duplicated.

    Parameters
    ----------
    path
        Filesystem location of the XLSX file. The sheet does not
        need to exist yet for writes.
    sheet
        Name of the sheet this handle scopes to.
    backend
        Default DataFrame backend for reads and writes.

    See Also
    --------
    Xlsx : Workbook-level counterpart.
    pandas.read_excel : Underlying pandas reader.
    pandas.ExcelWriter : Underlying pandas writer.
    mayutils.interfaces.filetypes.DataFile : Abstract tabular file
        interface implemented here.

    Examples
    --------
    >>> import tempfile
    >>> import pandas as pd
    >>> from pathlib import Path
    >>> with tempfile.TemporaryDirectory() as tmp:
    ...     path = Path(tmp) / "report.xlsx"
    ...     pd.DataFrame({"a": [1, 2, 3]}).to_excel(path, sheet_name="Summary", index=False)
    ...     sheet = XlsxSheet(path, sheet="Summary")
    ...     name = sheet.sheet
    ...     frame = sheet.read()
    >>> name
    'Summary'
    >>> frame.shape
    (3, 1)
    """

    suffix: ClassVar[str] = ".xlsx"

    def __init__(
        self,
        path: Path | str,
        /,
        *,
        sheet: str,
        backend: Backend[DataFrameType] | None = None,
    ) -> None:
        """
        Bind the handle to a ``(path, sheet)`` pair with a DataFrame backend.

        Delegates suffix validation to :class:`DataFile`'s initialiser
        and then records the target sheet name so per-sheet reads and
        writes can resolve the correct tab. No IO happens at
        construction time; the workbook is only touched when
        :meth:`_read`, :meth:`_write`, :meth:`schema`,
        :meth:`row_count`, or :meth:`iter_chunks` runs.

        Parameters
        ----------
        path
            XLSX file path.
        sheet
            Target sheet name.
        backend
            Default DataFrame backend.

        See Also
        --------
        Xlsx.sheet : Factory that constructs an :class:`XlsxSheet`
            from a workbook handle.
        mayutils.interfaces.filetypes.DataFile.__init__ : Delegated
            suffix validation that enforces the ``.xlsx`` suffix and
            raises :class:`ValueError` otherwise.

        Examples
        --------
        >>> sheet = XlsxSheet("report.xlsx", sheet="Summary")
        >>> sheet.sheet
        'Summary'
        """
        super().__init__(path, backend=backend)
        self.sheet = sheet

    def _identity(
        self,
    ) -> str:
        """
        Return the identity header including the bound sheet name.

        Overrides :meth:`DataFile._identity` so :meth:`__repr__`
        makes the sheet scope visible in the preview header. Used by
        both the plain-text and HTML repr paths inherited from
        :class:`DataFile` so Jupyter output clearly distinguishes
        between handles that share a workbook path but differ by
        sheet.

        Returns
        -------
            ``"XlsxSheet(<path>, sheet=<sheet>, backend=<backend>)"``.

        See Also
        --------
        Xlsx._identity : Workbook-level analogue.
        mayutils.interfaces.filetypes.DataFile._identity : Default
            implementation overridden here.

        Examples
        --------
        >>> XlsxSheet("report.xlsx", sheet="Summary")._identity()
        "XlsxSheet(report.xlsx, sheet='Summary', backend='pandas')"
        """
        return f"XlsxSheet({self.path!s}, sheet={self.sheet!r}, backend={self.backend.name!r})"

    def read(
        self,
        **kwargs: Any,  # noqa: ANN401
    ) -> DataFrameType:
        """
        Read the scoped sheet into a DataFrame of the requested backend.

        Dispatches to :func:`polars.read_excel` when
        :attr:`backend` is ``"polars"`` so polars callers get a
        native DataFrame without a pandas intermediate, and falls back
        to :func:`pandas.read_excel` otherwise. Both branches pin
        ``sheet_name`` to :attr:`sheet` so callers cannot accidentally
        override the scoped sheet via ``kwargs``.

        Parameters
        ----------
        **kwargs
            Forwarded to the backend-specific reader. The sheet name
            is fixed to :attr:`sheet` by this method.

        Returns
        -------
            Fully loaded DataFrame whose concrete type matches
            :attr:`backend`.

        See Also
        --------
        XlsxSheet._write : Inverse operation.
        pandas.read_excel : Backing pandas reader.
        polars.read_excel : Backing polars reader.
        Xlsx.read : Workbook-level counterpart that returns every
            sheet at once.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     pd.DataFrame({"a": [1, 2]}).to_excel(path, sheet_name="Summary", index=False)
        ...     sheet = XlsxSheet(path, sheet="Summary")
        ...     frame = sheet.read()
        >>> frame.shape
        (2, 1)
        >>> frame.columns.tolist()
        ['a']
        """
        if self.backend.name == "polars":
            return self.backend.cast(
                pl.read_excel(
                    source=self.path,
                    sheet_name=self.sheet,
                    **kwargs,
                )
            )

        return self.backend.cast(
            pd.read_excel(  # pyright: ignore[reportUnknownMemberType]
                io=self.path,
                sheet_name=self.sheet,
                **kwargs,
            )
        )

    def write(
        self,
        df: DataFrameType,
        /,
        **kwargs: Any,  # noqa: ANN401
    ) -> Self:
        """
        Serialise a DataFrame to the scoped sheet, merging into existing files.

        If the workbook already exists, the target sheet is
        replaced; any other sheets in the workbook are preserved. If
        the workbook does not yet exist, a new one is created with
        the scoped sheet as its only content. When the target is a
        brand-new workbook and ``df`` is a polars DataFrame, the
        write goes natively through
        :meth:`polars.DataFrame.write_excel`, avoiding a pandas
        round-trip. The append-to-existing path has to funnel
        through :class:`pandas.ExcelWriter` regardless of backend,
        because polars' writer cannot merge into an existing
        workbook.

        Parameters
        ----------
        df
            DataFrame to persist; its runtime type has already been
            validated against :attr:`backend` by
            :meth:`DataFile.write`.
        **kwargs
            Forwarded to the underlying writer —
            :meth:`polars.DataFrame.write_excel` on the polars fast
            path, :meth:`pandas.DataFrame.to_excel` otherwise (for
            example ``index=False``).

        Returns
        -------
            ``self``, for method chaining.

        See Also
        --------
        XlsxSheet._read : Inverse operation.
        pandas.ExcelWriter : Backing writer for the merge path.
        polars.DataFrame.write_excel : Backing writer for the fast
            path.
        Xlsx.write : Workbook-level counterpart that writes every
            sheet at once.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "out.xlsx"
        ...     sheet = XlsxSheet(path, sheet="Summary")
        ...     _ = sheet.write(pd.DataFrame({"a": [1]}))
        ...     existed = path.is_file()
        ...     readback = sheet.read()
        >>> existed
        True
        >>> readback.shape
        (1, 1)
        """
        if not self.path.is_file():  # noqa: SIM102
            if self.backend.name == "polars":
                df.write_excel(  # pyright: ignore[reportCallIssue]
                    workbook=self.path,
                    worksheet=self.sheet,
                    **kwargs,
                )

                return self

        if isinstance(df, pl.DataFrame):  # noqa: SIM108
            pandas_frame = df.to_pandas()
        else:
            pandas_frame = df

        mode = "a" if self.path.is_file() else "w"
        writer_kwargs: dict[str, Any] = {"engine": "openpyxl", "mode": mode}
        if mode == "a":
            writer_kwargs["if_sheet_exists"] = "replace"

        index_kwarg = kwargs.pop("index", False)

        with ExcelWriter(path=self.path, **writer_kwargs) as writer:  # pyright: ignore[reportUnknownVariableType]
            pandas_frame.to_excel(  # pyright: ignore[reportUnknownMemberType]
                excel_writer=writer,
                sheet_name=self.sheet,
                index=index_kwarg,
                **kwargs,
            )

        return self

    def excel_range(
        self,
        *,
        rows: int,
        columns: int,
        start_row: int = 1,
        start_column: int = 1,
    ) -> str:
        """
        Return an A1-style range notation scoped to :attr:`sheet`.

        Uses :meth:`Xlsx.column_to_excel` to translate column ordinals
        into spreadsheet letters, so ``excel_range(rows=10,
        columns=3)`` returns ``"Sheet1!A1:C10"`` for a sheet named
        ``"Sheet1"``. Useful when composing arguments to Google
        Sheets / Excel APIs that expect a bounded range reference.

        Parameters
        ----------
        rows
            Number of rows in the range (inclusive of ``start_row``).
        columns
            Number of columns in the range (inclusive of
            ``start_column``).
        start_row
            One-indexed row of the top-left cell.
        start_column
            One-indexed column of the top-left cell.

        Returns
        -------
            ``"<sheet>!<col1><row1>:<col2><row2>"``.

        See Also
        --------
        Xlsx.column_to_excel : Column ordinal to letter translation
            used internally.
        openpyxl.utils.cell.range_boundaries : Inverse helper that
            parses A1 ranges.

        Examples
        --------
        >>> XlsxSheet("report.xlsx", sheet="Sheet1").excel_range(
        ...     rows=10,
        ...     columns=3,
        ... )
        'Sheet1!A1:C10'
        >>> XlsxSheet("report.xlsx", sheet="Sheet1").excel_range(
        ...     rows=5,
        ...     columns=2,
        ...     start_row=2,
        ...     start_column=2,
        ... )
        'Sheet1!B2:C6'
        """
        start_letter = Xlsx.column_to_excel(start_column)
        end_letter = Xlsx.column_to_excel(start_column + columns - 1)
        end_row = start_row + rows - 1

        return f"{self.sheet}!{start_letter}{start_row}:{end_letter}{end_row}"

    def schema(
        self,
        *,
        sample_rows: int = DEFAULT_SCHEMA_SAMPLE_ROWS,
    ) -> dict[str, object]:
        """
        Infer the column-to-dtype mapping from the header and a sample.

        Reads the first ``sample_rows`` rows of the scoped sheet via
        :func:`pandas.read_excel` and captures the inferred dtypes
        per column. A sample is used rather than a full read because
        XLSX files lack cheap schema introspection and reading every
        row just to look at dtypes would be wasteful for large
        sheets.

        Parameters
        ----------
        sample_rows
            Number of rows to sample for dtype inference.

        Returns
        -------
            Column name to pandas dtype.

        See Also
        --------
        XlsxSheet.row_count : Companion helper that returns the total
            row count without reading the data.
        pandas.read_excel : Backing reader used for the sample.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     pd.DataFrame({"a": [1, 2], "b": ["x", "y"]}).to_excel(
        ...         path,
        ...         sheet_name="Summary",
        ...         index=False,
        ...     )
        ...     schema = XlsxSheet(path, sheet="Summary").schema()
        >>> sorted(schema)
        ['a', 'b']
        >>> str(schema["a"])
        'int64'
        >>> str(schema["b"])
        'object'
        """
        sample = pd.read_excel(  # pyright: ignore[reportUnknownMemberType]
            io=self.path,
            sheet_name=self.sheet,
            nrows=sample_rows,
        )

        return {column: sample[column].dtype for column in sample.columns}

    def row_count(
        self,
    ) -> int:
        """
        Return the row count via openpyxl's cached dimensions.

        Opens the workbook in read-only, data-only mode so the call
        is cheap and does not materialise cell values. Subtracts one
        from the raw ``max_row`` to account for the header row that
        pandas strips during a normal read, giving callers the count
        of data rows they would actually see.

        Returns
        -------
            Non-negative number of data rows (header row excluded).

        See Also
        --------
        openpyxl.worksheet.worksheet.Worksheet.max_row : Backing
            property used here.
        XlsxSheet.schema : Companion helper that returns dtypes.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     pd.DataFrame({"a": list(range(5))}).to_excel(
        ...         path,
        ...         sheet_name="Summary",
        ...         index=False,
        ...     )
        ...     count = XlsxSheet(path, sheet="Summary").row_count()
        >>> count
        5
        """
        workbook = openpyxl.load_workbook(
            filename=self.path,
            read_only=True,
            data_only=True,
        )

        try:
            worksheet = workbook[self.sheet]
            return (worksheet.max_row or 0) - 1
        finally:
            workbook.close()

    def iter_chunks(
        self,
        chunk_size: int,
        /,
        **kwargs: Any,  # noqa: ANN401
    ) -> Iterator[DataFrameType]:
        """
        Stream the sheet as DataFrame chunks of ``chunk_size`` rows.

        The full sheet is loaded once (openpyxl does not expose
        random-access chunked reads that preserve pandas' parsing
        semantics) and then sliced into row windows. The polars path
        uses :meth:`polars.DataFrame.slice` for zero-copy windows;
        the pandas path uses :meth:`pandas.DataFrame.iloc` slicing
        which shares underlying numpy buffers where possible.

        Parameters
        ----------
        chunk_size
            Upper bound on the number of rows per yielded chunk.
        **kwargs
            Forwarded to :func:`pandas.read_excel`.

        Yields
        ------
            Successive chunks whose concatenation equals
            :meth:`read`.

        See Also
        --------
        XlsxSheet._read : One-shot reader used internally to prime
            the stream.
        pandas.read_excel : Backing pandas reader.
        polars.read_excel : Backing polars reader.

        Examples
        --------
        >>> import tempfile
        >>> import pandas as pd
        >>> from pathlib import Path
        >>> with tempfile.TemporaryDirectory() as tmp:
        ...     path = Path(tmp) / "report.xlsx"
        ...     pd.DataFrame({"a": list(range(5))}).to_excel(
        ...         path,
        ...         sheet_name="Summary",
        ...         index=False,
        ...     )
        ...     sheet = XlsxSheet(path, sheet="Summary")
        ...     shapes = [chunk.shape for chunk in sheet.iter_chunks(2)]
        >>> shapes
        [(2, 1), (2, 1), (1, 1)]
        """
        backend = self.backend.name

        if backend == "polars":
            frame = pl.read_excel(
                source=self.path,
                sheet_name=self.sheet,
                **kwargs,
            )

            for start in range(0, frame.height, chunk_size):
                yield self.backend.cast(frame.slice(offset=start, length=chunk_size))

            return

        pandas_frame = pd.read_excel(  # pyright: ignore[reportUnknownMemberType]
            io=self.path,
            sheet_name=self.sheet,
            **kwargs,
        )

        for start in range(0, len(pandas_frame), chunk_size):
            yield self.backend.cast(pandas_frame.iloc[start : start + chunk_size])


__all__ = [
    "Xlsx",
    "XlsxSheet",
]
